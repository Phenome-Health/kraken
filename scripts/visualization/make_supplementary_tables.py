#!/usr/bin/env python3
"""
Build the KRAKEN supplementary-tables workbook (.xlsx) directly from a metagraph
JSON. A README/cover sheet (graph version, summary stats, notes) plus four simple
two-column, count-sorted tables:

  S1. Primary knowledge sources  (infores CURIE, edge count)    <- primary_knowledge_sources
  S2. Vocabulary prefixes        (prefix, node count)           <- node_prefixes
  S3. Node types                 (Biolink category, node count) <- node_categories
  S4. Edge types                 (Biolink predicate, edge count)<- edge_predicates

Usage:
    python make_supplementary_tables.py METAGRAPH.json -o supplementary_tables.xlsx
"""
import argparse
import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

SCRIPT_DIR = Path(__file__).resolve().parent

MULTI_NOTE = (
    "Note: these counts sum to MORE than the total number of KRAKEN nodes, because a "
    "single node can carry multiple {}. See the README sheet."
)

TABLES = [
    dict(
        title="S1. Primary knowledge sources",
        key="primary_knowledge_sources",
        kh="Primary knowledge source (infores CURIE)",
        vh="Edge count",
        desc="Every primary knowledge source and the number of edges attributed to it.",
        note=None,
    ),
    dict(
        title="S2. Vocabulary prefixes",
        key="node_prefixes",
        kh="Vocabulary prefix",
        vh="Node count",
        desc="Every identifier prefix and the number of node identifiers that use it.",
        note=MULTI_NOTE.format("equivalent identifiers (and thus multiple prefixes)"),
    ),
    dict(
        title="S3. Node types",
        key="node_categories",
        kh="Node type (Biolink category)",
        vh="Node count",
        desc="Every node type and the number of nodes assigned to it.",
        note=MULTI_NOTE.format("Biolink categories"),
    ),
    dict(
        title="S4. Edge types",
        key="edge_predicates",
        kh="Edge type (Biolink predicate)",
        vh="Edge count",
        desc="Every edge type and the number of edges of that type.",
        note=None,
    ),
]

README_NOTES = [
    "In S2 (prefixes) and S3 (node types), the counts sum to MORE than the total number of nodes: "
    "a single node can carry multiple equivalent identifiers (hence multiple vocabulary prefixes) "
    "and multiple Biolink categories.",
    "In S1 (primary knowledge sources) and S4 (edge types), the counts sum to the total number of "
    "edges: each edge has exactly one primary knowledge source and one predicate.",
]

# (metagraph summary key, README label) -- only those present are shown
SUMMARY_ROWS = [
    ("total_nodes", "Total nodes"),
    ("total_edges", "Total edges"),
    ("unique_node_categories", "Unique node types (Biolink categories)"),
    ("unique_node_prefixes", "Unique vocabulary prefixes"),
    ("unique_edge_predicates", "Unique edge types (Biolink predicates)"),
    ("unique_primary_knowledge_sources", "Unique primary knowledge sources"),
    ("unique_aggregator_knowledge_sources", "Unique aggregator knowledge sources"),
    ("unique_supporting_data_sources", "Unique supporting data sources"),
]


def _count_cell(cell):
    cell.number_format = "#,##0"
    cell.alignment = Alignment(horizontal="right")


def write_table(ws, t, data: dict):
    row = 1
    if t["note"]:  # short caveat above the header
        ws.cell(row, 1, t["note"])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws.cell(row, 1)
        c.font = Font(italic=True, color="595959")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 28
        row += 1

    ws.cell(row, 1, t["kh"]).font = Font(bold=True)
    ws.cell(row, 2, t["vh"]).font = Font(bold=True)
    ws.freeze_panes = ws.cell(row + 1, 1).coordinate

    max_key = len(t["kh"])
    for name, count in sorted(data.items(), key=lambda kv: -kv[1]):
        row += 1
        ws.cell(row, 1, name)
        _count_cell(ws.cell(row, 2, count))
        max_key = max(max_key, len(str(name)))

    ws.column_dimensions["A"].width = min(max_key + 2, 60)
    ws.column_dimensions["B"].width = max(len(t["vh"]) + 2, 12)


def build_readme(wb, meta, source_name):
    ws = wb.create_sheet("README", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 40
    r = 1

    def title(text, size):
        nonlocal r
        ws.cell(r, 1, text).font = Font(bold=True, size=size)
        r += 1

    def kv(label, value, count=False):
        nonlocal r
        ws.cell(r, 1, label)
        cell = ws.cell(r, 2, value)
        if count:
            _count_cell(cell)
        r += 1

    def note(text):
        nonlocal r
        ws.cell(r, 1, "• " + text)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 42
        r += 1

    title("KRAKEN knowledge graph — supplementary tables", 14)
    r += 1
    kv("Graph name", meta.get("graph", "—"))
    kv("Graph version", meta.get("version", "—"))
    kv("Biolink Model version", meta.get("biolink_version", "—"))
    kv("Source metagraph", source_name)
    kv("Generated", date.today().isoformat())
    r += 1

    title("Summary statistics", 12)
    summ = meta.get("summary", {})
    for key, label in SUMMARY_ROWS:
        if key in summ:
            kv(label, summ[key], count=True)
    r += 1

    title("Contents", 12)
    for t in TABLES:
        ws.cell(r, 1, t["title"]).font = Font(bold=True)
        ws.cell(r, 2, t["desc"]).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 1

    title("Notes", 12)
    for n in README_NOTES:
        note(n)


def make_pdf(pdf_path, meta, source_name):
    """Render the same content as a paginated, print-ready PDF (title page + tables)."""
    import textwrap

    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    plt.rcParams["pdf.fonttype"] = 42  # embed TrueType, not Type 3
    plt.rcParams["font.family"] = "sans-serif"
    plt.rcParams["font.sans-serif"] = ["Helvetica Neue", "Helvetica", "Arial", "DejaVu Sans"]
    PAGE = (8.5, 11.0)  # US Letter portrait
    ROWS_PER_PAGE = 44

    with PdfPages(pdf_path) as pdf:
        # --- title / README page ---
        fig = plt.figure(figsize=PAGE)
        y = [0.95]

        def line(label, val=None, size=10, bold=False, wrap=None):
            if wrap:
                for wl in textwrap.wrap(label, wrap):
                    fig.text(0.08, y[0], wl, fontsize=size)
                    y[0] -= 0.020
                y[0] -= 0.006
                return
            fig.text(0.08, y[0], label, fontsize=size, fontweight="bold" if bold else "normal")
            if val is not None:
                fig.text(0.62, y[0], val, fontsize=size)
            y[0] -= 0.030

        fig.text(0.08, y[0], "KRAKEN knowledge graph — supplementary tables", fontsize=15, fontweight="bold")
        y[0] -= 0.055
        line("Graph name", str(meta.get("graph", "—")))
        line("Graph version", str(meta.get("version", "—")))
        line("Biolink Model version", str(meta.get("biolink_version", "—")))
        line("Source metagraph", source_name)
        line("Generated", date.today().isoformat())
        y[0] -= 0.02
        line("Summary statistics", size=12, bold=True)
        summ = meta.get("summary", {})
        for key, label in SUMMARY_ROWS:
            if key in summ:
                line(label, f"{summ[key]:,}")
        y[0] -= 0.02
        line("Contents", size=12, bold=True)
        for t in TABLES:
            fig.text(0.08, y[0], t["title"], fontsize=9.5, fontweight="bold")
            fig.text(0.40, y[0], t["desc"], fontsize=9)
            y[0] -= 0.030
        y[0] -= 0.02
        line("Notes", size=12, bold=True)
        for n in README_NOTES:
            line("• " + n, size=9, wrap=105)
        pdf.savefig(fig)
        plt.close(fig)

        # --- one (paginated) section per table ---
        for t in TABLES:
            items = sorted(meta[t["key"]].items(), key=lambda kv: -kv[1])
            chunks = [items[i : i + ROWS_PER_PAGE] for i in range(0, len(items), ROWS_PER_PAGE)] or [[]]
            for pi, chunk in enumerate(chunks):
                fig = plt.figure(figsize=PAGE)
                title = t["title"] + ("  (continued)" if pi else "")
                fig.text(0.08, 0.955, title, fontsize=13, fontweight="bold")
                yh = 0.915
                fig.text(0.08, yh, t["kh"], fontsize=8.5, fontweight="bold")
                fig.text(0.93, yh, t["vh"], fontsize=8.5, fontweight="bold", ha="right")
                fig.add_artist(
                    plt.Line2D(
                        [0.08, 0.93], [yh - 0.008, yh - 0.008], color="#999999", lw=0.6, transform=fig.transFigure
                    )
                )
                dy = 0.855 / ROWS_PER_PAGE
                for ri, (name, count) in enumerate(chunk):
                    yy = yh - 0.022 - ri * dy
                    fig.text(0.08, yy, str(name), fontsize=7.5)
                    fig.text(0.93, yy, f"{count:,}", fontsize=7.5, ha="right")
                fig.text(
                    0.5,
                    0.03,
                    f"{t['title']} · page {pi + 1} of {len(chunks)}",
                    ha="center",
                    fontsize=7,
                    color="#888888",
                )
                pdf.savefig(fig)
                plt.close(fig)

    print(f"Wrote {pdf_path}")


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("metagraph", type=Path, help="Path to metagraph JSON")
    ap.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("supplementary_tables.xlsx"),
        help="Output .xlsx path (default beside this script)",
    )
    ap.add_argument(
        "--pdf",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Also write a paginated PDF alongside the .xlsx (default: yes; --no-pdf to skip)",
    )
    args = ap.parse_args()
    if not args.output.is_absolute():
        args.output = SCRIPT_DIR / args.output

    meta = json.loads(args.metagraph.read_text())
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    build_readme(wb, meta, args.metagraph.name)
    for t in TABLES:
        if t["key"] not in meta:
            raise SystemExit(f"metagraph is missing '{t['key']}' (needed for {t['title']})")
        ws = wb.create_sheet(title=t["title"])
        write_table(ws, t, meta[t["key"]])
        print(f"{t['title']}: {len(meta[t['key']]):,} rows")

    wb.save(args.output)
    print(f"Wrote {args.output}")

    if args.pdf:
        make_pdf(args.output.with_suffix(".pdf"), meta, args.metagraph.name)


if __name__ == "__main__":
    main()
