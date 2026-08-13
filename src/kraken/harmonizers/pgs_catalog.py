# pgs_catalog.py
import gzip
import heapq
import logging
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
import requests

from kraken.biolink_client import BiolinkClient
from kraken.harmonizers.base import BaseHarmonizer
from kraken.harmonizers.pgs_gene_annotator import ENSEMBL_GTF_URL, GeneAnnotator
from kraken.utils.constants import COMPUTATIONAL_MODEL, PGS_CATALOG_SOURCE_ID, STATISTICAL_ASSOCIATION
from kraken.utils.kg_io import save_to_jsonl

# --- v1 selection knobs ---
# Rather than a high global evaluation threshold (which silently drops important-but-less-studied traits like
# asthma/Alzheimer), we keep the best-validated score PER TRAIT. This keeps every trait's flagship PGS while
# cutting redundancy (e.g. ~22 CAD scores -> the single most-evaluated one).
MIN_EVALUATION_SAMPLE_SETS = 5  # a PGS must be validated in >= this many distinct evaluation sample sets
TOP_N_PER_TRAIT = 1  # keep only the N best-validated (most-evaluated) scores per mapped trait
# Per PGS, only its most-influential variants (by variance contribution, weight^2 * 2*AF*(1-AF)) are used, to
# bound work/edges. Two separate caps: gene identification aggregates variants -> genes (so the gene-edge count
# is bounded by the number of distinct genes regardless of how many variants feed in) and can afford a much
# larger pool; PGS->variant edges are one-per-variant, so they keep the tighter cap. Both draw from the same
# variance-ranked ordering (the gene pool is a superset of the variant-edge pool).
TOP_N_VARIANTS_FOR_GENES = 100_000
TOP_N_VARIANTS_FOR_EDGES = 10_000
# A PGS linking to ~every gene is uninformative, so per PGS we keep only its top genes, ranked by the summed
# variance contribution of each gene's variants (the same variance measure used to rank variants -- more
# principled than summed |effect_weight|, since it accounts for allele frequency). Small scores keep all their
# genes (fewer than this anyway); genome-wide scores keep only their most-influential genes.
TOP_N_GENES_PER_PGS = 500

# --- PLACEHOLDER Biolink types (pending a types review; each is a single swappable constant) ---
# There is no clean Biolink class for a polygenic score; InformationContentEntity is a stand-in.
PGS_NODE_CATEGORY = "biolink:InformationContentEntity"  # TODO(types): revisit (ClinicalFinding? a custom PGS class?)
TRAIT_STUB_CATEGORY = "biolink:NamedThing"  # minimal stub; real category arrives when ontology sources merge in
GENE_CATEGORY = "biolink:Gene"  # gene endpoints; emitted as ENSEMBL:<ENSG>, merge into graph gene nodes via ER
VARIANT_CATEGORY = "biolink:SequenceVariant"  # variant endpoints (rsID-bearing variants only)
PGS_TRAIT_PREDICATE = "biolink:positively_correlated_with"  # a PGS statistically correlates with its trait
PGS_GENE_PREDICATE = "biolink:related_to"  # TODO(types): revisit -- semantics are "involves a variant located in"
PGS_VARIANT_PREDICATE = "biolink:related_to"  # TODO(types): revisit -- semantics are "has scoring component variant"

PGS_CURIE_PREFIX = "PGS"  # PGS Catalog accession is e.g. "PGS000013"; emitted as "PGS:000013" (an id we mint)
# Gene/variant endpoint ids come from external sources (Ensembl GTF, scoring-file rsIDs), so we normalize them
# to biolink-compliant CURIEs via biomapper2 rather than assuming a prefix. These are the biomapper2 vocab names.
GENE_VOCAB = "ensembl"
VARIANT_VOCAB = "dbsnp"

# Per-PGS harmonized scoring file (GRCh38 positions). Cached alongside the metadata file (input_file.parent).
HARMONIZED_SCORING_URL = (
    "https://ftp.ebi.ac.uk/pub/databases/spot/pgs/scores/{pgs_id}/ScoringFiles/Harmonized/{pgs_id}_hmPOS_GRCh38.txt.gz"
)


class PGSCatalogHarmonizer(BaseHarmonizer):
    """Harmonizer for the PGS Catalog metadata bundle (``pgs_all_metadata.xlsx``).

    v1 scope: from the ~7k scores, keep the single best-validated PGS per mapped trait (ranked by number of
    distinct evaluation sample sets, gated at >= ``MIN_EVALUATION_SAMPLE_SETS``). Each selected score becomes a
    PGS node linked to (a) its mapped EFO/MONDO trait(s), (b) the genes its top variants fall in, and (c) those
    top variants themselves when they carry an rsID. The gene/variant links come from downloading each PGS's
    harmonized (GRCh38) scoring file, taking its top variants by variance contribution (weight^2 * 2*AF*(1-AF))
    -- a larger pool for genes (``TOP_N_VARIANTS_FOR_GENES``) than for variant edges (``TOP_N_VARIANTS_FOR_EDGES``,
    which are one-per-variant) -- and (for genes) overlapping their positions against an Ensembl gene model (see
    ``pgs_gene_annotator``) -- a high-confidence "variant located in gene" relationship (overlap only, no
    nearest-gene), not a causal one. PGS->variant edges are only emitted for variants with an rsID (so they merge
    cleanly with graph variant nodes). Minimal stub nodes are minted for trait/gene/variant endpoints so nothing
    orphans. Scoring files and the gene model are cached under the metadata file's directory (``input_file.parent``).

    All node/edge Biolink types are placeholders -- see the module-level ``*_CATEGORY`` / ``*_PREDICATE``
    constants (pending a types review); each is a single swappable constant.
    """

    source_infores = PGS_CATALOG_SOURCE_ID

    def __init__(self, biolink_client: BiolinkClient):
        super().__init__(biolink_client)
        self._curie_cache: dict[tuple[str, str], list[str]] = {}  # (vocab, local id) -> normalized curie(s)

    def harmonize(
        self,
        nodes_output: Path,
        edges_output: Path,
        *,
        input_file: Path | None = None,
        nodes_input: Path | None = None,
        edges_input: Path | None = None,
    ):
        if not input_file:
            raise ValueError(f"{self.source_name} requires input_file")
        logging.info(f"Harmonizing {self.source_name}: {input_file} -> {nodes_output}, {edges_output}")

        wb = openpyxl.load_workbook(input_file, read_only=True)
        scores = self._load_scores(wb)  # pgs_id -> metadata dict
        evaluations = self._load_evaluations(wb)  # pgs_id -> set of distinct evaluation sample sets
        wb.close()

        selected = self._select_top_per_trait(scores, evaluations)

        cache_dir = input_file.parent  # scoring files + gene model live alongside the metadata file
        annotator = self._load_gene_annotator(cache_dir)

        # Stream each PGS's nodes/edges to disk as we go, rather than accumulating everything in memory (a
        # genome-wide score's gene/variant edges add up). Stub endpoints (trait/gene/variant) are deduped across
        # PGS via lightweight seen-sets of CURIEs, so each shared node is written only once.
        save_to_jsonl([], nodes_output, mode="w")  # truncate any prior output; we append per-PGS below
        save_to_jsonl([], edges_output, mode="w")
        written_traits: set[str] = set()
        written_genes: set[str] = set()
        written_variants: set[str] = set()
        trait_edge_count = gene_edge_count = variant_edge_count = 0
        for i, pgs_id in enumerate(sorted(selected), 1):
            meta = scores[pgs_id]
            node_batch: list[dict] = [self._make_pgs_node(pgs_id, meta)]
            edge_batch: list[dict] = []
            for trait_curie, trait_label in meta["trait_curies"]:
                if trait_curie not in written_traits:
                    written_traits.add(trait_curie)
                    node_batch.append(self._make_trait_stub(trait_curie, trait_label))
                edge_batch.append(self._make_pgs_trait_edge(pgs_id, trait_curie, len(evaluations.get(pgs_id, ()))))
                trait_edge_count += 1
            gene_edges, variant_edges = self._scoring_edges_for_pgs(
                pgs_id, cache_dir, annotator, node_batch, written_genes, written_variants
            )
            edge_batch.extend(gene_edges)
            edge_batch.extend(variant_edges)
            gene_edge_count += len(gene_edges)
            variant_edge_count += len(variant_edges)

            save_to_jsonl(node_batch, nodes_output, mode="a")
            save_to_jsonl(edge_batch, edges_output, mode="a")
            if i % 25 == 0:
                logging.info(f"  ...processed {i}/{len(selected)} PGS")

        logging.info(
            f"{self.source_name} harmonization complete: {len(selected)} PGS nodes, {len(written_traits)} trait / "
            f"{len(written_genes)} gene / {len(written_variants)} variant stubs; {trait_edge_count} PGS->trait + "
            f"{gene_edge_count} PGS->gene + {variant_edge_count} PGS->variant edges "
            f"(selected {len(selected)} of {len(scores)} scores)"
        )

    # ------------------------------------------------------------------ loading

    def _load_scores(self, wb) -> dict[str, dict[str, Any]]:
        scores: dict[str, dict] = {}
        for r in wb["Scores"].iter_rows(min_row=2, values_only=True):
            pgs_id = r[0]
            if not pgs_id:
                continue
            scores[pgs_id] = {
                "name": r[1],
                "reported_trait": r[2],
                "mapped_trait_label": r[3],
                "mapped_trait_efo": r[4],
                "trait_curies": self._parse_trait_curies(r[3], r[4]),
                "development_method": r[5],
                "genome_build": r[7],
                "num_variants": self._to_int(r[8]),
                "weight_type": r[10],
                "pgp_id": r[11],
                "pmid": r[12],
                "ftp_link": r[18],  # scoring file -- retained for the later gene-annotation increment
                "release_date": str(r[19]) if r[19] else None,  # str(): cell may be a datetime (not JSON-safe)
            }
        return scores

    def _load_evaluations(self, wb) -> dict[str, set]:
        """pgs_id -> set of distinct evaluation sample sets (PSS). Distinct sample sets, not raw metric rows,
        is the honest 'independently validated in N cohorts' measure (one cohort often reports several metrics)."""
        sample_sets: dict[str, set] = defaultdict(set)
        for r in wb["Performance Metrics"].iter_rows(min_row=2, values_only=True):
            pgs_id, sample_set = r[1], r[2]
            if pgs_id and sample_set:
                sample_sets[pgs_id].add(sample_set)
        return sample_sets

    # ------------------------------------------------------------------ selection

    def _select_top_per_trait(self, scores: dict, evaluations: dict[str, set]) -> set[str]:
        eval_count = lambda pgs_id: len(evaluations.get(pgs_id, ()))
        # Group eligible scores by each mapped trait, then keep the top-N most-evaluated per trait.
        by_trait: dict[str, list[str]] = defaultdict(list)
        for pgs_id, meta in scores.items():
            if eval_count(pgs_id) >= MIN_EVALUATION_SAMPLE_SETS:
                for trait_curie, _ in meta["trait_curies"]:
                    by_trait[trait_curie].append(pgs_id)
        selected: set[str] = set()
        for pgs_ids in by_trait.values():
            selected.update(sorted(pgs_ids, key=eval_count, reverse=True)[:TOP_N_PER_TRAIT])
        return selected

    # ------------------------------------------------------------------ node/edge builders

    def _make_pgs_node(self, pgs_id: str, meta: dict) -> dict:
        curie = self._pgs_curie(pgs_id)
        publications = [f"PMID:{meta['pmid']}"] if meta.get("pmid") else None
        attributes = {
            "pgs_catalog_id": pgs_id,  # native accession, e.g. "PGS000013"
            "reported_trait": meta["reported_trait"],
            "mapped_trait_label": meta["mapped_trait_label"],
            "mapped_trait_efo": meta["mapped_trait_efo"],
            "pgs_development_method": meta["development_method"],
            "genome_build": meta["genome_build"],
            "number_of_variants": meta["num_variants"],
            "type_of_variant_weight": meta["weight_type"],
            "pgs_publication_id": meta["pgp_id"],
            "scoring_file_ftp": meta["ftp_link"],
            "release_date": meta["release_date"],
        }
        return self.create_node(
            curie=curie,
            categories=[PGS_NODE_CATEGORY],
            provided_by=self.source_infores,
            equivalent_ids=[curie],
            name=meta["name"],
            description=f"Polygenic score for {meta['reported_trait']}" if meta.get("reported_trait") else None,
            publications=publications,
            attributes={k: v for k, v in attributes.items() if v not in (None, "")},
        )

    def _make_trait_stub(self, trait_curie: str, trait_label: str | None) -> dict:
        return self.create_node(
            curie=trait_curie,
            categories=[TRAIT_STUB_CATEGORY],
            provided_by=self.source_infores,
            equivalent_ids=[trait_curie],
            name=trait_label,
        )

    def _make_pgs_trait_edge(self, pgs_id: str, trait_curie: str, num_sample_sets: int) -> dict:
        return self.create_edge(
            subject_id=self._pgs_curie(pgs_id),
            object_id=trait_curie,
            predicate=PGS_TRAIT_PREDICATE,
            primary_ks=self.source_infores,
            knowledge_level=STATISTICAL_ASSOCIATION,
            agent_type=COMPUTATIONAL_MODEL,
            attributes={"num_evaluation_sample_sets": num_sample_sets} if num_sample_sets else None,
        )

    # ------------------------------------------------------------------ gene linking

    def _load_gene_annotator(self, cache_dir: Path) -> GeneAnnotator:
        gtf_path = cache_dir / "gene_model" / ENSEMBL_GTF_URL.rsplit("/", 1)[1]
        if not self._download(ENSEMBL_GTF_URL, gtf_path):
            raise RuntimeError(f"Could not download the gene model from {ENSEMBL_GTF_URL}")
        return GeneAnnotator(gtf_path)

    def _scoring_edges_for_pgs(
        self,
        pgs_id: str,
        cache_dir: Path,
        annotator: GeneAnnotator,
        node_batch: list[dict],
        written_genes: set[str],
        written_variants: set[str],
    ) -> tuple[list[dict], list[dict]]:
        """From a PGS's top variants, build (a) PGS->gene edges (variants overlapping a protein-coding gene body --
        high-confidence "located in", no nearest-gene fallback), and (b) PGS->variant edges for the variants that
        carry an rsID (so they merge cleanly with graph variant nodes). Newly-seen gene/variant stub nodes are
        appended to `node_batch` (deduped across PGS via `written_genes`/`written_variants`). Returns
        (gene_edges, variant_edges)."""
        scoring_path = self._download_scoring_file(pgs_id, cache_dir)
        if not scoring_path:
            logging.warning(f"No harmonized GRCh38 scoring file for {pgs_id}; skipping its gene/variant edges.")
            return [], []

        # Variance-ranked (desc) top variants; genes use the whole pool, variant edges only its top slice.
        top_variants = self._top_variants(scoring_path, TOP_N_VARIANTS_FOR_GENES)
        gene_hits: dict[str, list] = (
            {}
        )  # ensembl_gene_id -> [num_variants_in_gene, summed_variance_contribution, symbol]
        variant_edges: list[dict] = []
        for rank, (chrom, pos, rsid, variance_contribution) in enumerate(top_variants):
            # PGS -> gene: overlap with a gene body only (high confidence)
            for ensg, symbol in annotator.genes_at(chrom, pos):
                hit = gene_hits.get(ensg)
                if hit:
                    hit[0] += 1
                    hit[1] += variance_contribution
                else:
                    gene_hits[ensg] = [1, variance_contribution, symbol]
            # PGS -> variant: only among the top TOP_N_VARIANTS_FOR_EDGES, and only when an rsID is available.
            # Normalize the rsID to a biolink-compliant CURIE (don't assume the prefix); skip if it won't resolve.
            if rank < TOP_N_VARIANTS_FOR_EDGES and rsid and rsid.startswith("rs"):
                variant_curie = next(iter(self._normalize_curies(VARIANT_VOCAB, rsid)), None)
                if variant_curie:
                    if variant_curie not in written_variants:
                        written_variants.add(variant_curie)
                        node_batch.append(self._make_variant_stub(variant_curie))
                    variant_edges.append(self._make_pgs_variant_edge(pgs_id, variant_curie))

        # Keep only the top genes by summed variance contribution (avoid linking a PGS to ~every gene)
        top_genes = heapq.nlargest(TOP_N_GENES_PER_PGS, gene_hits.items(), key=lambda item: item[1][1])
        gene_edges = []
        for ensg, (count, _variance_sum, symbol) in top_genes:
            # Normalize the Ensembl gene id to a biolink-compliant CURIE (don't assume the prefix)
            gene_curie = next(iter(self._normalize_curies(GENE_VOCAB, ensg)), None)
            if not gene_curie:
                continue
            if gene_curie not in written_genes:
                written_genes.add(gene_curie)
                node_batch.append(self._make_gene_stub(gene_curie, symbol))
            gene_edges.append(self._make_pgs_gene_edge(pgs_id, gene_curie, count))
        return gene_edges, variant_edges

    def _top_variants(self, scoring_path: Path, limit: int) -> list[tuple[str, int, str | None, float]]:
        """Return the top `limit` variants of a harmonized scoring file as (chrom, pos, rsid, variance_contribution),
        ranked (descending) by variance contribution (weight^2 * 2*AF*(1-AF)). AF (allelefrequency_effect) is in the
        files; when it is missing/invalid we assume a common variant (AF=0.5). The raw effect weight is used only to
        compute the (internal) ranking -- it is not returned or published (some PGS licenses restrict it). Streamed
        via a bounded heap -- O(limit) memory even for genome-wide scores."""

        def scored_rows():
            with gzip.open(scoring_path, "rt") as f:
                header = None
                for line in f:
                    if line.startswith("#"):
                        continue
                    cols = line.rstrip("\n").split("\t")
                    if header is None:
                        header = {c: i for i, c in enumerate(cols)}
                        continue
                    chrom = self._col(cols, header, "hm_chr") or self._col(cols, header, "chr_name")
                    pos = self._col(cols, header, "hm_pos") or self._col(cols, header, "chr_position")
                    weight = self._to_float(self._col(cols, header, "effect_weight"))
                    if not chrom or not pos or weight is None:
                        continue
                    try:
                        position = int(pos)
                    except ValueError:
                        continue
                    af = self._to_float(self._col(cols, header, "allelefrequency_effect"))
                    if af is None or not (0 < af < 1):
                        af = 0.5  # assume a common variant when the effect-allele frequency is unavailable
                    variance_contribution = weight * weight * 2 * af * (1 - af)
                    rsid = self._col(cols, header, "hm_rsID") or self._col(cols, header, "rsID")
                    yield (variance_contribution, chrom, position, rsid, weight)

        top = heapq.nlargest(limit, scored_rows(), key=lambda row: row[0])
        return [(chrom, pos, rsid, variance) for (variance, chrom, pos, rsid, _weight) in top]

    def _download_scoring_file(self, pgs_id: str, cache_dir: Path) -> Path | None:
        dest = cache_dir / "scoring_files" / f"{pgs_id}_hmPOS_GRCh38.txt.gz"
        return self._download(HARMONIZED_SCORING_URL.format(pgs_id=pgs_id), dest)

    @staticmethod
    def _download(url: str, dest: Path) -> Path | None:
        """Download `url` to `dest` (cached: returns immediately if it already exists). Writes to a .part file
        and renames on completion so a partial download is never cached as complete. Returns None on 404."""
        if dest.exists():
            return dest
        dest.parent.mkdir(parents=True, exist_ok=True)
        logging.info(f"Downloading {url}")
        with requests.get(url, stream=True, timeout=120) as resp:
            if resp.status_code == 404:
                return None
            resp.raise_for_status()
            part = dest.with_name(dest.name + ".part")
            with open(part, "wb") as fh:
                for chunk in resp.iter_content(chunk_size=1 << 20):
                    fh.write(chunk)
            part.rename(dest)
        return dest

    def _make_gene_stub(self, gene_curie: str, symbol: str | None) -> dict:
        return self.create_node(
            curie=gene_curie,
            categories=[GENE_CATEGORY],
            provided_by=self.source_infores,
            equivalent_ids=[gene_curie],
            name=symbol,
        )

    def _make_pgs_gene_edge(self, pgs_id: str, gene_curie: str, num_variants: int) -> dict:
        # NOTE: we intentionally do NOT publish weight-derived values (e.g. summed variance contribution): they
        # can back-derive a variant's effect weight (esp. for single-variant genes), and some PGS have licenses
        # that restrict redistributing scoring-file values. Weights are used internally (ranking) only; only the
        # (license-safe) variant count is emitted. The scoring-file link on the PGS node is the path to weights.
        return self.create_edge(
            subject_id=self._pgs_curie(pgs_id),
            object_id=gene_curie,
            predicate=PGS_GENE_PREDICATE,
            primary_ks=self.source_infores,
            knowledge_level=STATISTICAL_ASSOCIATION,
            agent_type=COMPUTATIONAL_MODEL,
            attributes={"num_scoring_variants_in_gene": num_variants},
        )

    def _make_variant_stub(self, variant_curie: str) -> dict:
        return self.create_node(
            curie=variant_curie,
            categories=[VARIANT_CATEGORY],
            provided_by=self.source_infores,
            equivalent_ids=[variant_curie],
        )

    def _make_pgs_variant_edge(self, pgs_id: str, variant_curie: str) -> dict:
        # NOTE: the variant's effect weight is intentionally NOT published -- it is a verbatim scoring-file value
        # and some PGS licenses restrict redistributing those. The edge asserts membership only; the scoring-file
        # link on the PGS node is the (license-respecting) path to the weights.
        return self.create_edge(
            subject_id=self._pgs_curie(pgs_id),
            object_id=variant_curie,
            predicate=PGS_VARIANT_PREDICATE,
            primary_ks=self.source_infores,
            knowledge_level=STATISTICAL_ASSOCIATION,
            agent_type=COMPUTATIONAL_MODEL,
        )

    @staticmethod
    def _col(cols: list[str], header: dict[str, int], name: str) -> str | None:
        i = header.get(name)
        if i is None or i >= len(cols):
            return None
        value = cols[i].strip()
        return value or None

    # ------------------------------------------------------------------ helpers

    @staticmethod
    def _pgs_curie(pgs_id: str) -> str:
        return f"{PGS_CURIE_PREFIX}:{pgs_id.removeprefix('PGS')}"  # PGS000013 -> PGS:000013

    def _parse_trait_curies(self, label: str | None, efo: str | None) -> list[tuple[str, str | None]]:
        """Map the mapped-trait id(s) to (normalized_curie, label) pairs. The ids come from an external source
        (PGS Catalog), so rather than trust the raw 'PREFIX_local' form we route each through biomapper2's
        normalizer to validate and canonicalize the CURIE. Both columns may hold multiple '|'/','-separated
        values; labels are zipped positionally when counts line up."""
        if not efo:
            return []
        raw_ids = [e.strip() for e in re.split(r"[|,]", str(efo)) if e.strip()]
        labels = [l.strip() for l in re.split(r"[|,]", str(label))] if label else []
        pairs = []
        for i, raw in enumerate(raw_ids):
            label_i = labels[i] if i < len(labels) else (label or None)
            for curie in self._normalize_trait_id(raw):
                pairs.append((curie, label_i))
        return pairs

    def _normalize_trait_id(self, raw: str) -> list[str]:
        """Split a raw mapped-trait id ('MONDO_0005010') into vocab + local id and normalize via biomapper2."""
        sep = "_" if "_" in raw else ":"
        prefix, _, local = raw.partition(sep)  # 'MONDO_0005010' -> ('MONDO', '0005010')
        return self._normalize_curies(prefix, local) if prefix and local else []

    def _normalize_curies(self, vocab: str, local_id: str) -> list[str]:
        """Return biolink-compliant CURIE(s) for `vocab` + `local_id` via biomapper2 -- canonicalizing the prefix
        and validating the id format, rather than assuming a prefix. Used for every endpoint whose id comes from
        an external source (traits, genes, variants); ids we mint ourselves (PGS:...) are left as-is. Cached;
        warns once (then drops) when an id can't be resolved."""
        key = (vocab, local_id)
        if key in self._curie_cache:
            return self._curie_cache[key]
        resolved, _, _ = self.normalizer.get_curies({vocab: local_id}, stop_on_invalid_id=False, log_warnings=False)
        curies = list(resolved)
        if not curies:
            logging.warning(f"{self.source_name}: could not normalize id {vocab}:{local_id}; dropping it.")
        self._curie_cache[key] = curies
        return curies

    @staticmethod
    def _to_int(value: Any) -> int | None:
        try:
            return int(str(value).replace(",", "")) if value not in (None, "") else None
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _to_float(value: Any) -> float | None:
        try:
            return float(value) if value not in (None, "") else None
        except (ValueError, TypeError):
            return None
